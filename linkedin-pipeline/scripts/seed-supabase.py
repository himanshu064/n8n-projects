"""
Seed the linkedin_running_order table in Supabase from the Excel file.

This is the script shortcut for STEPS.md Workflow 1 — same result, no clicking.
Safe to re-run: rows whose scheduled_date already exists are skipped
(upsert with ignore-duplicates), so it never creates duplicates.

Setup (once):
    set SUPABASE_URL=https://<your-project>.supabase.co
    set SUPABASE_SERVICE_ROLE_KEY=<service-role-key>
  (PowerShell:  $env:SUPABASE_URL="..." ; $env:SUPABASE_SERVICE_ROLE_KEY="...")

Usage:
    python seed-supabase.py [path-to-xlsx]

    Default file: running-order-sample.xlsx next to this script
    (run generate-running-order.py first if it doesn't exist).

Uses only the Python standard library + openpyxl (no supabase package needed).
"""
import json
import os
import sys
import urllib.request
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

EXPECTED_HEADER = ["scheduled_date", "process_name", "style", "sequence_number"]


def die(msg: str) -> None:
    print(f"ERROR: {msg}")
    sys.exit(1)


def load_rows(xlsx_path: Path) -> list[dict]:
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        die("Excel file is empty")

    header = [str(h).strip().lower() if h else "" for h in rows[0]]
    if header[: len(EXPECTED_HEADER)] != EXPECTED_HEADER:
        die(f"Expected columns {EXPECTED_HEADER}, got {header}")

    out = []
    for i, row in enumerate(rows[1:], start=2):
        if all(v is None for v in row):
            continue  # skip blank trailing rows
        d, process, style, seq = row[:4]
        if not d or not process or not style or seq is None:
            die(f"Row {i}: empty required field: {row}")
        d_iso = d.date().isoformat() if hasattr(d, "date") else str(d)
        date.fromisoformat(d_iso)  # validates format, raises if bad
        out.append({
            "scheduled_date": d_iso,
            "process_name": str(process).strip(),
            "style": str(style).strip(),
            "sequence_number": int(seq),
        })
    return out


def validate(rows: list[dict]) -> None:
    dates = [r["scheduled_date"] for r in rows]
    dupes = {d for d in dates if dates.count(d) > 1}
    if dupes:
        die(f"Duplicate dates: {sorted(dupes)}")
    if len(rows) != 90:
        print(f"WARNING: expected 90 rows, found {len(rows)} — continuing anyway")


def seed(rows: list[dict], url: str, key: str) -> None:
    req = urllib.request.Request(
        f"{url.rstrip('/')}/rest/v1/linkedin_running_order"
        "?on_conflict=scheduled_date",
        data=json.dumps(rows).encode(),
        method="POST",
        headers={
            "apikey": key,
            "Authorization": f"Bearer {key}",
            "Content-Type": "application/json",
            # insert new dates, silently skip dates already in the table
            "Prefer": "resolution=ignore-duplicates,return=representation",
        },
    )
    try:
        with urllib.request.urlopen(req) as resp:
            inserted = json.loads(resp.read() or b"[]")
    except urllib.error.HTTPError as e:
        die(f"Supabase returned {e.code}: {e.read().decode()[:500]}")
    print(f"Inserted {len(inserted)} new rows "
          f"({len(rows) - len(inserted)} already existed, skipped)")


def main() -> None:
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        die("Set SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY environment "
            "variables first (see header of this file)")

    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else \
        Path(__file__).parent / "running-order-sample.xlsx"
    if not xlsx.exists():
        die(f"{xlsx} not found — run generate-running-order.py first, "
            "or pass the path to the client's Excel file")

    rows = load_rows(xlsx)
    validate(rows)
    print(f"Seeding {len(rows)} rows from {xlsx.name} ...")
    seed(rows, url, key)


if __name__ == "__main__":
    main()
